
# https://openpyxl.readthedocs.io/en/stable/
# https://learn.microsoft.com/en-us/sql/connect/python/pyodbc/step-3-proof-of-concept-connecting-to-sql-using-pyodbc?view=sql-server-ver16
import pandas as pd
import pyodbc
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np
import re

database_username = 'test_sql_user'
database_password = 'testHrw4772'
database_server = 'SERVER8'
database_name = 'DWH'

cnxn = pyodbc.connect(
    'Driver={SQL Server Native Client 11.0};SERVER=' + database_server +
    ';DATABASE=' + database_name +
    ';UID=' + database_username +
    ';PWD=' + database_password)

def db_conn_pd():
    sel = "SELECT * FROM [DWH].[dbo].[Sales]"
    df = pd.read_sql(sel, cnxn)
    print(type(df))

def db_conn_odbc():
    cursor = cnxn.cursor()
    #sel = cursor.execute("SELECT  [ТаблицаЗатраты0].[НаименованиеДокументаПлатежаПолное],  [ДокументыДДС_К2].[Документ] from [ТаблицаЗатраты0], [ДокументыДДС_К2]  where [dbo1].[ТаблицаЗатраты0].[НаименованиеДокументаПлатежаПолное] =  [ДокументыДДС_К2].[Документ]")
    sel = cursor.execute('select * from [ДокументыДДС_К2]')
    df = pd.DataFrame(sel)
    df.to_clipboard(index=False, header=True)
    print(df)

def df_file():
    wb = load_workbook('затраты_с2018.xlsx')
    sheet = wb.active
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0]
    df = df.drop(0, axis=0)
    pd.set_option('display.max_columns', None)
    df['Дата затрат'] = pd.to_datetime(df['Дата затрат'])
    file = open("текст.txt", "w")
    for i in range(1, df.shape[0]):
        if type(df['Дата платежа'][i]) == str:
            if df['Договор контрагента'][i] is not None:
                file.write(f"{i}, {df['Документ платежа'][i]}, {df.loc[i, 'Договор контрагента']} , {df['Договор контрагента'].astype(str)[i] + ' ' + df['Дата платежа'][i]}" )
                df.loc[i, 'Договор контрагента'] = df['Договор контрагента'].astype(str)[i] + ' ' + df['Дата платежа'][i]
               # file.write(df.loc[i, 'Договор контрагента'])
            else:
                df.loc[i, 'Договор контрагента'] = df['Дата платежа'][i]
                #file.write(df.loc[i, 'Договор контрагента'])
                file.write(f"{i}, {df['Документ платежа'][i]}, {df.loc[i, 'Договор контрагента']} , {df['Дата платежа'][i]}" )
            df.loc[i, 'Дата платежа'] = df['Дата затрат'][i]
            file.write(df.loc[i, 'Дата платежа'])
                #f"{i}, {df['Документ платежа'][i]}, {df.loc[i, 'Дата платежа']} , {df['Договор контрагента'].astype(str)[i] + ' ' + df['Дата платежа'][i]}")
            if np.isnat(np.datetime64(str(df['Дата затрат'][i]))):
                df.loc[i, 'Дата платежа'] = str(df['Год затрат'][i]) + "-" + str(df['Месяц затрат'][i]) + "-" + "01"
                file.write(df.loc[i, 'Дата платежа'])
        # elif df['Дата платежа'][i] is None:
        #     df.loc[i, 'Дата платежа'] = df['Дата затрат'][i]
        #     if np.isnat(np.datetime64(str(df['Дата затрат'][i]))):
        #          df.loc[i, 'Дата платежа'] = str(df['Год затрат'][i]) + "-" + str(df['Месяц затрат'][i]) + "-" + "01"
        # elif np.isnat(np.datetime64(str(df['Дата затрат'][i]))) or df['Дата затрат'][i] is None:
        #     df.loc[i, 'Дата затрат'] = re.findall(r'\d\d\.\d\d\.\d{4}', str(df['Документ платежа'][i]))[0]
        if len(str(df['Комментарий'][i])) > 256:
            df.loc[i, 'Комментарий'] = (df['Комментарий'][i])[:255]
            file.write(df.loc[i, 'Комментарий'])
        if df['Сумма платежа'][i] is None:
            df.loc[i, 'Сумма платежа'] = df['Сумма затрат'][i]
            file.write(df.loc[i, 'Сумма платежа'])
        if df['Сумма затрат'][i] is None:
            df.loc[i, 'Сумма затрат'] = df['Сумма платежа'][i]
            file.write(df.loc[i, 'Сумма затрат'])
        df['Дата платежа'] = pd.to_datetime(str(df['Дата платежа'][i]))
        df['Дата затрат'] = pd.to_datetime(str(df['Дата затрат'][i]))
        df = df.astype({'Сумма платежа': 'float', 'Сумма затрат': 'float', 'Год затрат': 'int', 'Месяц затрат': 'int',
                        'Квартал затрат': 'int'})
    #print(df.iloc[[28337]], df.iloc[[18651]], df.iloc[[18460]], df[[18459]])
    # print(df['Дата платежа'][i], df['Дата затрат'][i],  df['Договор контрагента'][i])
    #print(df.dtypes, df.isnull().sum())
    file.close()
    return df


def df_insert(df):
    cursor = cnxn.cursor()
    for index, row in df.iterrows():
        rows = row['Дата платежа'], row['подразделение'], row['Сумма платежа'], row['Сумма затрат'], row['Дата затрат'], row['Вид операции'], row['Документ платежа'], row['Комментарий'], row['Укрупненная сатья ДДС'], row['Укрупненная сатья Затрат'], row['Контрагент'], row['Договор контрагента'], row['Статья затрат'], row['Организация'], row['Статья ДДС'], row['Месяц затрат'], row['Квартал затрат'], row['Год затрат'], row['ЗАО']
        try: cursor.execute(f"""INSERT INTO [DWH].[dbo].[ТаблицаЗатраты0]  ([ДатаПлатежа], [Подразделение], [СуммаПлатежа], [СуммаЗатрат], [ДатаЗатрат], [ВидОперации],
            [НаименованиеДокументаПлатежаПолное], [Комментарий], [СтатьяЗатратSql], [УкрупненнаяСтатьяSql], [НаименованиеКонтрагент], [НаименованиеДоговорКонтрагента],
            [Отдел], [НаименованиеОрганизация], [СтатьяДДСsql], [МесяцЗатрат], [КварталЗатрат], [ГодЗатрат], [ЗАО])   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", rows)
        except Exception as e:
            print(row['Дата платежа'], row['Сумма платежа'], index, e)
    cnxn.commit()

from openpyxl.utils.dataframe import dataframe_to_rows


def s_file():
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df_file()):
        ws.append(r)
    ws.delete_rows(2, 1)
    wb.save("pandas_openpyxl.xlsx")

#def p_file():


def db_update(col, num):
    cursor = cnxn.cursor()
    cursor.execute(f"""UPDATE [ТаблицаЗатраты0]
SET [ТаблицаЗатраты0].[{col}] = (
SELECT b.[{col}]
FROM [ЭлементарныеПартии].[dbo].[ДокументыДДС_К2] b
WHERE b.[Документ] = [ТаблицаЗатраты0].[НаименованиеДокументаПлатежаПолное])
WHERE [ТаблицаЗатраты0].[{col}] is null AND [ТаблицаЗатраты0].[ДатаПлатежа] LIKE '2018-10-{num}'""")
    cnxn.commit()



def main():
  #  db_update('ТипДокумента', '29')
    #df_file()
    df_insert(df_file())
    #db_conn_odbc()
   # r_file()


if __name__ == '__main__':
    main()


    # def db_connection(cnxn):
    #     database_username = 'test_sql_user'
    #     database_password = 'testHrw4772'
    #     database_server = 'SERVER8'
    #     # database_server = 'tcp:SERVER8.ЭлементарныеПартии.windows.net'
    #     # database_server = f"SERVER8 (SERVER8\vmokrietsova)"

# openpyxl
# https://openpyxl.readthedocs.io/en/stable/
# pyodbc
# https://learn.microsoft.com/en-us/sql/connect/python/pyodbc/step-3-proof-of-concept-connecting-to-sql-using-pyodbc?view=sql-server-ver16


# self.cnxn = pyodbc.connect(
#     "Driver={SQL Server Native Client 11.0};"
#     f"Server={config['server']};"
#     f"Database={config['database']};"
#     f"Uid={config['user']};"
#     f"Pwd={config['password']};"
# )
#   cnxn = pyodbc.connect(
#       'Driver={SQL Server Native Client 11.0};SERVER=' + database_server +
#       ';DATABASE=' + database_name +
#       ';UID=' + database_username +
#       ';PWD=' + database_password)
#   cursor = cnxn.cursor()

   # wb = load_workbook(filename=BytesIO(input_excel.read()))
    #book = openpyxl.open('Соответствие_ДДС.xlsx', read_only = True)
    #print(book)
    # for row in range(1, sheet.max_row+1):
    #     data_plat = sheet[row][4].value
    #     podrazdel = sheet[row][12].value
    #     #zao = sheet[row][20].value
    #     sum_plat = sheet[row][15].value
    #     data_zatr = sheet[row][6].value
    #     vid_oper = sheet[row][9].value
    #     dokument_plat = sheet[row][5].value
    #     komment = sheet[row][11].value
    #     ukrup_stat_dds = sheet[row][13].value
    #     ukrup_stat_zatrat = sheet[row][14].value
    #     sum_zatrat = sheet[row][16].value
    #     kontagent = sheet[row][2].value
    #     dogovor_kontrag = sheet[row][3].value
    #     statia_zatrat = sheet[row][7].value
    #     org = sheet[row][1].value
    #     statia_dds = sheet[row][8].value
    #cursor = cnxn.cursor()
 #    cursor.execute(f"""INSERT INTO [DWH].[dbo].[ТаблицаЗатраты0]  ([ДатаПлатежа], [Подразделение], [СуммаПлатежа], [СуммаЗатрат], [ДатаЗатрат], [ВидОперации],
 # [НаименованиеДокументаПлатежаПолное], [Комментарий], [СтатьяЗатратSql], [УкрупненнаяСтатьяSql], [НаименованиеКонтрагент], [НаименованиеДоговорКонтрагента],
 #  [Отдел], [НаименованиеОрганизация], [СтатьяДДСsql])   VALUES ('{data_plat}', '{podrazdel}', '{sum_plat}', '{sum_zatrat}', '{data_zatr}', '{vid_oper}',
 #  '{dokument_plat}', '{komment}', '{ukrup_stat_dds}', '{ukrup_stat_zatrat}', '{kontagent}', '{dogovor_kontrag}', '{statia_zatrat}', '{org}', '{statia_dds}' )""")
 #    #cnxn.commit()
        #print(df)
        #return df
        #print(data_plat, data_zatr)


